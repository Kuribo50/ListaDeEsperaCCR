from datetime import date
from io import BytesIO

from django.core.files.base import ContentFile
from django.db.models import Q, F
from django.http import HttpResponse
from django.db import transaction
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from rest_framework import status
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.pacientes.models import Paciente
from apps.usuarios.permissions import IsAdminOrAdministrativoRole

from .models import ImportacionMensual
from .parser import parsear_derivaciones, previsualizar_derivaciones
from .serializers import ImportacionDerivacionesSerializer


MESES_SHEET = {
    "ENERO": 1,
    "FEBRERO": 2,
    "MARZO": 3,
    "ABRIL": 4,
    "MAYO": 5,
    "JUNIO": 6,
    "JULIO": 7,
    "AGOSTO": 8,
    "SEPTIEMBRE": 9,
    "OCTUBRE": 10,
    "NOVIEMBRE": 11,
    "DICIEMBRE": 12,
}

MESES_LABEL = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Septiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre",
}


def _bool_from_request(value) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "on", "si"}


def _mes_y_anio_referencia(importacion: ImportacionMensual) -> tuple[int, int]:
    return (
        importacion.mes_datos or importacion.mes,
        importacion.anio_datos or importacion.anio,
    )


def _periodo_q(mes: int, anio: int) -> Q:
    return Q(mes_datos=mes, anio_datos=anio) | Q(
        mes_datos__isnull=True,
        anio_datos__isnull=True,
        mes=mes,
        anio=anio,
    )


def _serialize_importacion(importacion: ImportacionMensual) -> dict:
    mes_ref, anio_ref = _mes_y_anio_referencia(importacion)
    return {
        "id": importacion.id,
        "archivo_nombre": importacion.archivo_nombre or importacion.archivo.name.rsplit("/", 1)[-1],
        "mes": importacion.mes,
        "anio": importacion.anio,
        "mes_datos": importacion.mes_datos,
        "anio_datos": importacion.anio_datos,
        "mes_label": MESES_LABEL.get(mes_ref, str(mes_ref)),
        "periodo_label": f"{MESES_LABEL.get(mes_ref, mes_ref)} {anio_ref}",
        "usuario_id": importacion.usuario_id,
        "usuario_nombre": importacion.usuario.nombre if importacion.usuario else None,
        "fecha_subida": importacion.fecha_subida.isoformat(),
        "estado": importacion.estado,
        "estado_label": importacion.get_estado_display(),
        "total_registros": importacion.total_registros,
        "registros_importados": importacion.registros_importados,
        "duplicados": importacion.duplicados,
        "errores": importacion.errores,
        "reemplazada_por": importacion.reemplazada_por_id,
    }


class ImportarDerivacionesView(APIView):
    parser_classes = [MultiPartParser]
    permission_classes = [IsAdminOrAdministrativoRole]

    def post(self, request):
        serializer = ImportacionDerivacionesSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        archivo = serializer.validated_data["archivo"]
        hoy = date.today()
        anio_datos = serializer.validated_data.get("anio", hoy.year)
        resultado = parsear_derivaciones(archivo)
        meses_detectados = resultado["meses_detectados"]
        forzar_reemplazo = _bool_from_request(request.data.get("forzar_reemplazo", False))
        modo_suplementar = _bool_from_request(request.data.get("modo_suplementar", False))

        conflictos = []
        importaciones_previas: dict[int, ImportacionMensual] = {}
        for nombre_hoja, count in meses_detectados.items():
            mes_num = MESES_SHEET.get(nombre_hoja.upper())
            if not mes_num or count <= 0:
                continue
            importacion_previa = (
                ImportacionMensual.objects.filter(
                    _periodo_q(mes_num, anio_datos),
                    estado__in=[
                        ImportacionMensual.Estado.COMPLETADO,
                        ImportacionMensual.Estado.CON_ERRORES,
                    ],
                )
                .order_by("-fecha_subida")
                .first()
            )
            if importacion_previa:
                importaciones_previas[mes_num] = importacion_previa
                conflictos.append(
                    {
                        "hoja": nombre_hoja,
                        "mes": mes_num,
                        "anio": anio_datos,
                        "importados_previos": importacion_previa.registros_importados,
                        "fecha_subida_previa": importacion_previa.fecha_subida.isoformat(),
                        "importacion_id": importacion_previa.id,
                    }
                )

        if conflictos and not forzar_reemplazo and not modo_suplementar:
            return Response(
                {
                    "tipo": "conflicto_mes",
                    "mensaje": "Ya existen datos importados para estos meses.",
                    "conflictos": conflictos,
                    "pregunta": "¿Desea reemplazar los registros existentes de estos meses?",
                },
                status=status.HTTP_409_CONFLICT,
            )

        with transaction.atomic():
            archivo.seek(0)
            archivo_bytes = archivo.read()
            archivo_nombre = archivo.name
            nuevas_importaciones: dict[int, ImportacionMensual] = {}

            for nombre_hoja, count in meses_detectados.items():
                mes_num = MESES_SHEET.get(nombre_hoja.upper())
                if not mes_num or count <= 0:
                    continue

                nueva_importacion = ImportacionMensual.objects.create(
                    archivo=ContentFile(archivo_bytes, name=archivo_nombre),
                    archivo_nombre=archivo_nombre,
                    mes=hoy.month,
                    anio=hoy.year,
                    mes_datos=mes_num,
                    anio_datos=anio_datos,
                    usuario=request.user,
                    estado=(
                        ImportacionMensual.Estado.COMPLETADO
                        if not resultado["errores"]
                        else ImportacionMensual.Estado.CON_ERRORES
                    ),
                    total_registros=resultado["total"],
                    registros_importados=resultado["importados"],
                    duplicados=resultado["duplicados"],
                    errores=resultado["errores"],
                )
                nuevas_importaciones[mes_num] = nueva_importacion

            if forzar_reemplazo and conflictos:
                for mes_num, importacion_previa in importaciones_previas.items():
                    importacion_previa.estado = ImportacionMensual.Estado.REEMPLAZADO
                    importacion_previa.reemplazada_por = nuevas_importaciones.get(mes_num)
                    importacion_previa.save(update_fields=["estado", "reemplazada_por"])

            pacientes_a_crear = resultado.get("pacientes", [])
            if pacientes_a_crear:
                import uuid as _uuid
                batch_prefix = _uuid.uuid4().hex[:8].upper()  # e.g. A3F1C2B7
                # Link each patient to the correct ImportacionMensual (by month)
                for i, p in enumerate(pacientes_a_crear):
                    mes_derivacion = p.fecha_derivacion.month
                    importacion_correspondiente = nuevas_importaciones.get(mes_derivacion)
                    if importacion_correspondiente:
                        p.importacion_origen = importacion_correspondiente
                    elif nuevas_importaciones:
                        p.importacion_origen = list(nuevas_importaciones.values())[0]
                    # Assign a guaranteed-unique temporary id_ccr to avoid UNIQUE conflicts
                    p.id_ccr = f"TMP-{batch_prefix}-{i + 1:06d}"

                creados = Paciente.objects.bulk_create(pacientes_a_crear, batch_size=200)
                for paciente in creados:
                    paciente.id_ccr = str(paciente.pk)
                Paciente.objects.bulk_update(creados, ["id_ccr"], batch_size=200)
                resultado["importados"] = len(creados)

            # LÓGICA DE RECURRENTES (Meses de espera)
            # Buscamos pacientes que ya existían y les sumamos 1 al contador de meses de espera
            # si están en un estado de lista de espera.
            
            # Obtenemos los RUTs de registros que fueron marcados como duplicados (ya existen)
            # Nota: el parser los marca como es_duplicado=True
            ruts_duplicados = set()
            if "registros" in resultado:
                for reg in resultado["registros"]:
                    if reg.get("es_duplicado") and reg.get("rut"):
                        ruts_duplicados.add(reg["rut"])
            
            if ruts_duplicados:
                # Incrementamos el contador para los que ya estaban en el sistema
                # pero SOLO si están en un estado de espera
                # y NO pertenecen ya a una importación de este mismo periodo (para evitar doble conteo por "partes")
                meses_periodo = [MESES_SHEET.get(m.upper()) for m in meses_detectados.keys() if MESES_SHEET.get(m.upper())]
                
                pats_a_actualizar = Paciente.objects.filter(
                    rut__in=ruts_duplicados,
                    estado__in=[Paciente.Estado.PENDIENTE, Paciente.Estado.RESCATE]
                ).exclude(
                    importacion_origen__mes_datos__in=meses_periodo,
                    importacion_origen__anio_datos=anio_datos
                )
                
                # Obtenemos datos para loggear en el historial antes de actualizar o durante
                pats_data = list(pats_a_actualizar.values_list('id', 'estado'))
                
                if pats_data:
                    # 1. Incrementamos contador
                    pats_a_actualizar.update(n_meses_espera=F('n_meses_espera') + 1)
                    
                    # 2. Registramos en el historial (MovimientoPaciente)
                    from apps.pacientes.models import MovimientoPaciente
                    meses_str = ", ".join([MESES_LABEL.get(m, str(m)) for m in meses_periodo])
                    
                    movimientos = [
                        MovimientoPaciente(
                            paciente_id=pid,
                            usuario=request.user,
                            estado_anterior=None, # Indicamos que no hay cambio de estado, sino actualización
                            estado_nuevo=estado,
                            notas=f"Registrado nuevamente en lista de espera: {meses_str} {anio_datos}"
                        ) for pid, estado in pats_data
                    ]
                    MovimientoPaciente.objects.bulk_create(movimientos, batch_size=200)

            # Update the success counts locally
            for importacion in nuevas_importaciones.values():
                importacion.registros_importados = resultado["importados"]
                importacion.save(update_fields=["registros_importados"])

            resultado.pop("pacientes", None)
            return Response(resultado, status=status.HTTP_201_CREATED)


class PrevisualizarDerivacionesView(APIView):
    parser_classes = [MultiPartParser]
    permission_classes = [IsAdminOrAdministrativoRole]

    def post(self, request):
        serializer = ImportacionDerivacionesSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        archivo = serializer.validated_data["archivo"]
        resultado = previsualizar_derivaciones(archivo)
        return Response(resultado, status=status.HTTP_200_OK)


class HistorialImportacionesView(APIView):
    permission_classes = [IsAdminOrAdministrativoRole]

    def get(self, request):
        historial = (
            ImportacionMensual.objects.select_related("usuario", "reemplazada_por")
            .all()
        )
        return Response([_serialize_importacion(item) for item in historial])


class HistorialImportacionesMesView(APIView):
    permission_classes = [IsAdminOrAdministrativoRole]

    def get(self, request, mes: int, anio: int):
        mes = int(mes)
        anio = int(anio)
        historial = (
            ImportacionMensual.objects.select_related("usuario", "reemplazada_por")
            .filter(_periodo_q(mes, anio))
            .order_by("-fecha_subida")
        )
        return Response(
            {
                "mes": mes,
                "anio": anio,
                "mes_label": MESES_LABEL.get(mes, str(mes)),
                "items": [_serialize_importacion(item) for item in historial],
            }
        )

    @transaction.atomic
    def delete(self, request, mes: int, anio: int):
        try:
            mes = int(mes)
            anio = int(anio)
        except (TypeError, ValueError):
            return Response(
                {"detail": "Periodo inválido."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if mes < 1 or mes > 12:
            return Response(
                {"detail": "Mes fuera de rango."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        historial_qs = ImportacionMensual.objects.filter(_periodo_q(mes, anio)).order_by(
            "-fecha_subida"
        )
        if not historial_qs.exists():
            return Response(
                {"detail": "No hay importaciones para ese periodo."},
                status=status.HTTP_404_NOT_FOUND,
            )

        importaciones = list(historial_qs)
        archivos_eliminados = 0
        pacientes_eliminados = 0
        
        for importacion in importaciones:
            # Solo eliminamos pacientes de esta importación que:
            # 1. No tengan kinesiólogo asignado
            # 2. Estén en estado PENDIENTE o RESCATE
            pats_a_borrar = importacion.pacientes_creados.filter(
                kine_asignado__isnull=True,
                estado__in=[Paciente.Estado.PENDIENTE, Paciente.Estado.RESCATE]
            )
            pacientes_eliminados += pats_a_borrar.count()
            pats_a_borrar.delete()
            
            # Los pacientes que ya están asignados o en otro estado (INGRESADO, etc.)
            # NO se borran. Al borrar la importación abajo, su campo 'importacion_origen'
            # pasará a ser NULL (gracias al on_delete=SET_NULL en el modelo).
            
            if importacion.archivo:
                importacion.archivo.delete(save=False)
                archivos_eliminados += 1
            importacion.delete()

        return Response(
            {
                "mes": mes,
                "anio": anio,
                "pacientes_eliminados": pacientes_eliminados,
                "importaciones_eliminadas": len(importaciones),
                "archivos_eliminados": archivos_eliminados,
            },
            status=status.HTTP_200_OK,
        )


class PlantillaImportacionView(APIView):
    permission_classes = [IsAdminOrAdministrativoRole]

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "ENERO"

        ws.merge_cells("A1:I1")
        ws["A1"] = "Derivación a Centro Comunitario de Rehabilitación Cesfam Dr. Alberto Reyes"
        ws["A1"].font = Font(bold=True, size=11)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers_f2 = {
            "A2": "FECHA",
            "B2": "NOMBRE",
            "C2": "RUT",
            "D2": "EDAD",
            "E2": "DESDE",
            "F2": "DIAGNÓSTICO MÉDICO",
            "G2": "PROFESIONAL DERIVADO",
            "H2": "GRADO PRIORIDAD",
            "I2": "OBSERVACIONES",
        }
        fill_header = PatternFill("solid", fgColor="1B5E3B")
        font_header = Font(bold=True, color="FFFFFF", size=10)
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell_ref, val in headers_f2.items():
            c = ws[cell_ref]
            c.value = val
            c.font = font_header
            c.fill = fill_header
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
        ws.row_dimensions[2].height = 24

        ejemplos = {
            "A3": "DD/MM/YYYY",
            "B3": "NOMBRE COMPLETO",
            "C3": "12345678-9",
            "D3": "65",
            "E3": "CAR / HT / CST",
            "F3": "LUMBAGO",
            "G3": "KINESIOLOGO",
            "H3": "ALTA / MEDIANA / MODERADA",
            "I3": "Observaciones opcionales",
        }
        fill_ejemplo = PatternFill("solid", fgColor="E8F5EE")
        font_ejemplo = Font(italic=True, color="3A5A3A", size=9)
        for cell_ref, val in ejemplos.items():
            c = ws[cell_ref]
            c.value = val
            c.font = font_ejemplo
            c.fill = fill_ejemplo
            c.alignment = Alignment(horizontal="center")
            c.border = border

        datos_ejemplo = [
            (
                "15/01/2025",
                "JUAN PÉREZ GONZÁLEZ",
                "12345678-9",
                65,
                "CAR",
                "LUMBAGO",
                "KINESIOLOGO",
                "ALTA",
                "Dolor crónico lumbar",
            ),
            (
                "20/01/2025",
                "MARÍA SOTO RAMÍREZ",
                "9876543-2",
                72,
                "HT",
                "GONARTROSIS (GES)",
                "KINESIOLOGO",
                "MEDIANA",
                "",
            ),
            (
                "22/01/2025",
                "PEDRO MUÑOZ VEGA",
                "15432198-K",
                45,
                "CST",
                "HOMBRO DOLOROSO",
                "KINESIOLOGO",
                "MODERADA",
                "ECO compatible",
            ),
        ]
        for i, fila in enumerate(datos_ejemplo, start=4):
            for j, val in enumerate(fila, start=1):
                c = ws.cell(row=i, column=j, value=val)
                c.font = Font(size=10)
                c.border = border
                if j == 1:
                    c.number_format = "DD/MM/YYYY"

        anchos = [14, 30, 14, 7, 12, 28, 18, 22, 40]
        for i, ancho in enumerate(anchos, start=1):
            ws.column_dimensions[chr(64 + i)].width = ancho

        ws2 = wb.create_sheet("INSTRUCCIONES")
        instrucciones = [
            ("INSTRUCCIONES DE USO", True),
            ("", False),
            (
                "1. Esta planilla es el formato oficial para subir derivaciones al sistema ListaEsperaCCR.",
                False,
            ),
            (
                "2. Cada hoja representa un mes. Renombrar la hoja con el nombre del mes en mayúsculas:",
                False,
            ),
            (
                "   ENERO, FEBRERO, MARZO, ABRIL, MAYO, JUNIO, JULIO, AGOSTO, SEPTIEMBRE, OCTUBRE, NOVIEMBRE, DICIEMBRE",
                False,
            ),
            ("3. Los datos deben comenzar en la fila 3 (después de los headers).", False),
            ("", False),
            ("COLUMNAS OBLIGATORIAS:", True),
            ("  FECHA          → Formato DD/MM/YYYY", False),
            ("  NOMBRE         → Nombre completo del paciente", False),
            ("  RUT            → Sin puntos, con guión (ej: 12345678-9)", False),
            ("  EDAD           → Número entero", False),
            ("  DESDE          → Sigla del centro de origen", False),
            ("  DIAGNÓSTICO    → Diagnóstico médico", False),
            ("  GRADO PRIORIDAD→ ALTA, MEDIANA o MODERADA", False),
            ("", False),
            ("CENTROS VÁLIDOS EN COLUMNA DESDE:", True),
            (
                "  CAR, CST, CCEQ, CCE, CES, HT, TMT, FST, HLH, TMT HT, FST HT, CEQ, CESFAM, CECOSF, SANTO",
                False,
            ),
            ("", False),
            ("PRIORIDADES VÁLIDAS:", True),
            ("  ALTA (atención urgente)", False),
            ("  MEDIANA (atención normal)", False),
            ("  MODERADA (atención moderada)", False),
        ]
        for i, (texto, bold) in enumerate(instrucciones, start=1):
            c = ws2.cell(row=i, column=1, value=texto)
            c.font = Font(bold=bold, size=10, color="1B5E3B" if bold else "2A3A2A")
        ws2.column_dimensions["A"].width = 90

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="Plantilla_Derivaciones_CCR.xlsx"'
        return response


class ResetPoblacionView(APIView):
    """
    DELETE /api/importar/reset
    Elimina TODOS los pacientes sin kinesiolo'go asignado.
    Los pacientes ASIGNADOS se conservan intactos.
    Tambien limpia el historial de importaciones.
    Solo accesible para ADMIN.
    """
    permission_classes = [IsAdminOrAdministrativoRole]

    @transaction.atomic
    def delete(self, request):
        if not (request.user.is_authenticated and getattr(request.user, 'rol', None) == 'ADMIN'):
            return Response(
                {'detail': 'Solo el administrador puede resetear la poblacion.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        # Solo eliminamos pacientes que:
        # 1. No tengan kinesiólogo asignado
        # 2. Estén en estado PENDIENTE o RESCATE
        # (Los INGRESADOS, EGRESADOS, ALTA, DERIVADO, ABANDONO se mantienen)
        sin_asignar_qs = Paciente.objects.filter(
            kine_asignado__isnull=True,
            estado__in=[Paciente.Estado.PENDIENTE, Paciente.Estado.RESCATE]
        )
        total_eliminados = sin_asignar_qs.count()
        sin_asignar_qs.delete()

        importaciones_qs = ImportacionMensual.objects.all()
        total_importaciones = importaciones_qs.count()
        for imp in importaciones_qs:
            if imp.archivo:
                try:
                    imp.archivo.delete(save=False)
                except Exception:
                    pass
        importaciones_qs.delete()

        return Response(
            {
                'pacientes_eliminados': total_eliminados,
                'importaciones_eliminadas': total_importaciones,
                'mensaje': f'Se eliminaron {total_eliminados} pacientes sin asignar y {total_importaciones} registros de importacion.',
            },
            status=status.HTTP_200_OK,
        )
