from datetime import date, datetime
import re
import unicodedata
import uuid

from openpyxl import load_workbook

from apps.pacientes.models import Paciente
from apps.pacientes.services import categoria_por_diagnostico, prioridad_normalizada

CENTROS_VALIDOS = {
    "CAR", "CST", "CCEQ", "CCE", "CES", "HT", "TMT", "FST",
    "HLH", "TMT HT", "FST HT", "CEQ", "CESFAM", "CECOSF", "SANTO",
}

MESES_IGNORAR = {"LISTAS", "INSTRUCCIONES"}
MESES_VALIDOS = {
    "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
    "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE",
}

MESES_NUM = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4,
    "MAYO": 5, "JUNIO": 6, "JULIO": 7, "AGOSTO": 8,
    "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12,
}

# Alias normalizados para la detección dinámica de columnas
# IMPORTANTE: Las listas van de más específico a más genérico. Se usa coincidencia EXACTA.
COLUMN_ALIASES: dict[str, list[str]] = {
    "fecha":        ["FECHA DERIV", "FECHA DERIV.", "FECHA DERIVACION",
                     "F DERIVACION", "FECHA"],
    "nombre":       ["NOMBRE", "NOMBRE COMPLETO", "PACIENTE", "NOMBRES Y APELLIDOS"],
    "rut":          ["RUT", "RUT PACIENTE", "RUN"],
    "edad":         ["EDAD", "EDAD ANOS", "ANOS"],
    "desde":        ["PERCÁPITA / DESDE", "PERCAPITA / DESDE", "PERCAPITA/DESDE",
                     "PERCAPITA DESDE", "DESDE", "PERCAPITA", "CENTRO ORIGEN"],
    "diagnostico":  ["DIAGNÓSTICO", "DIAGNOSTICO MEDICO", "DIAGNOSTICO", "DX"],
    "profesional":  ["PROFESIONAL DERIVADO", "PROFESIONAL DERIVACION", "PROFESIONAL",
                     "PROF DERIVADO", "KINESIOLOGO"],
    "prioridad":    ["PRIORIDAD", "GRADO PRIORIDAD", "GRADO DE PRIORIDAD", "GRADO"],
    "mayor_60":     ["≥60", "≥ 60", ">= 60", ">=60", ">60", "60", "MAYOR 60", "MAYOR60"],
    "observaciones":["OBSERVACIONES", "OBJETIVOS DEL TRATAMIENTO OBSERVACIONES",
                     "OBJETIVOS DEL TRATAMIENTO", "OBS", "COMENTARIOS", "NOTAS"],
}


def _cs(valor) -> str:
    if valor is None:
        return ""
    return re.sub(r"[\s\t]+", " ", str(valor)).strip()


def _normalizar(valor: str) -> str:
    texto = _cs(valor).upper()
    sin_tildes = "".join(
        c for c in unicodedata.normalize("NFD", texto)
        if unicodedata.category(c) != "Mn"
    )
    # Remover puntuación al inicio/final para que "FECHA DERIV." == "FECHA DERIV"
    sin_puntuacion = re.sub(r"^[.\-_/\s]+|[.\-_/\s]+$", "", sin_tildes)
    return " ".join(sin_puntuacion.split())


def _parsear_fecha(valor):
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if valor is None:
        return None

    texto = str(valor).strip()
    # Buscar primero formato con año de 4 dígitos
    match4 = re.search(r"(\d{2}[./-]\d{2}[./-]\d{4}|\d{4}-\d{2}-\d{2})", texto)
    if match4:
        texto = match4.group(1)
        for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(texto, fmt).date()
            except ValueError:
                continue

    # Buscar formato con año de 2 dígitos (ej: 30-09-24)
    match2 = re.search(r"(\d{2}[./-]\d{2}[./-]\d{2})$", texto)
    if match2:
        texto2 = match2.group(1)
        for fmt in ("%d.%m.%y", "%d/%m/%y", "%d-%m-%y"):
            try:
                return datetime.strptime(texto2, fmt).date()
            except ValueError:
                continue

    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%y", "%d/%m/%y", "%d-%m-%y"):
        try:
            return datetime.strptime(texto, fmt).date()
        except ValueError:
            continue

    try:
        numero = float(texto)
        from openpyxl.utils.datetime import from_excel
        convertido = from_excel(numero)
        if isinstance(convertido, datetime):
            return convertido.date()
        if isinstance(convertido, date):
            return convertido
    except Exception:
        pass

    return None


def _mes_desde_hoja(sheet_name: str) -> str | None:
    nombre = _normalizar(sheet_name)
    for mes in MESES_VALIDOS:
        if re.search(rf"\b{re.escape(mes)}\b", nombre):
            return mes
    return None


def _detectar_columnas_dinamicas(ws) -> dict[str, int] | None:
    """
    Busca en las primeras 8 filas de la hoja una fila con encabezados reconocibles
    y devuelve un mapeo campo->número_de_columna (1-indexed).
    Usa coincidencia EXACTA con los aliases normalizados para evitar falsos positivos.
    Devuelve None si no puede identificar las columnas mínimas.
    """
    # Pre-normalizar todos los aliases
    aliases_norm: dict[str, list[str]] = {
        field: [_normalizar(a) for a in aliases]
        for field, aliases in COLUMN_ALIASES.items()
    }

    for row_idx in range(1, 9):
        row_vals = {}
        all_empty = True
        for col in range(1, 22):
            raw = ws.cell(row_idx, col).value
            if raw is not None:
                all_empty = False
            norm = _normalizar(_cs(raw))
            row_vals[col] = norm

        if all_empty:
            continue

        mapping: dict[str, int] = {}
        for field, norm_aliases in aliases_norm.items():
            for col, cell_norm in row_vals.items():
                if not cell_norm:
                    continue
                # Coincidencia EXACTA: la celda normalizada debe estar en los aliases
                if cell_norm in norm_aliases:
                    if field not in mapping:
                        mapping[field] = col
                    break

        # Necesitamos al menos nombre, rut y diagnóstico para considerar válida la hoja
        if all(k in mapping for k in ("nombre", "rut", "diagnostico")):
            return {"header_row": row_idx, **mapping}

    return None


def _dup_key(rut: str, fecha_derivacion: date, diagnostico: str) -> tuple[str, date, str]:
    return (rut, fecha_derivacion, (diagnostico or "").upper().strip())


def _fecha_preview(fecha_derivacion: date | None) -> str:
    return fecha_derivacion.strftime("%d/%m/%Y") if fecha_derivacion else ""


def _registro_preview(
    *,
    hoja: str,
    fila: int,
    nombre: str,
    rut: str,
    fecha_derivacion: date | None,
    edad: int,
    diagnostico: str,
    prioridad: str,
    percapita_desde: str,
    profesional: str,
    observaciones: str,
    mayor_60: bool,
    categoria: str,
    es_duplicado: bool = False,
    error: str = "",
) -> dict:
    if error:
        estado = "ERROR"
    elif es_duplicado:
        estado = "DUPLICADO"
    else:
        estado = "OK"

    return {
        "hoja": hoja,
        "fila": fila,
        "nombre": nombre,
        "rut": rut,
        "fecha_derivacion": _fecha_preview(fecha_derivacion),
        "edad": edad,
        "diagnostico": diagnostico,
        "prioridad": prioridad,
        "percapita_desde": percapita_desde,
        "profesional": profesional,
        "observaciones": observaciones,
        "mayor_60": mayor_60,
        "categoria": categoria,
        "es_duplicado": es_duplicado,
        "estado": estado,
        "error": error,
    }


def _procesar_hoja(ws, sheet_name: str, columnas: dict, header_row: int,
                   existentes: set, duplicados_vistos: set) -> dict:
    """
    Procesa filas de una hoja usando el mapeo dinámico de columnas.
    Retorna dict con registros, pacientes_a_crear, errores, total, duplicados, fechas.
    """
    registros = []
    pacientes_a_crear = []
    errores = []
    total = 0
    duplicados = 0
    fechas_hoja: list[date] = []

    for row_idx in range(header_row + 1, ws.max_row + 1):
        nombre = _cs(ws.cell(row_idx, columnas["nombre"]).value) if "nombre" in columnas else ""
        rut_raw = _cs(ws.cell(row_idx, columnas["rut"]).value) if "rut" in columnas else ""
        diagnostico_bruto = _cs(ws.cell(row_idx, columnas["diagnostico"]).value) if "diagnostico" in columnas else ""
        profesional_bruto = _cs(ws.cell(row_idx, columnas.get("profesional", 0)).value) if "profesional" in columnas else ""
        prioridad_bruta = _cs(ws.cell(row_idx, columnas.get("prioridad", 0)).value) if "prioridad" in columnas else ""
        observaciones_brutas = _cs(ws.cell(row_idx, columnas.get("observaciones", 0)).value) if "observaciones" in columnas else ""

        if not nombre or not rut_raw:
            continue

        # Saltar filas que son encabezados repetidos
        nombre_norm = _normalizar(nombre)
        if nombre_norm in {"NOMBRE", "N°", "NOMBRE COMPLETO", "PACIENTE"}:
            continue
        # Saltar si toda la fila de campos clave está vacía (suma de campos clínicos)
        if not diagnostico_bruto and not profesional_bruto and not prioridad_bruta and not observaciones_brutas:
            continue

        total += 1
        fecha_derivacion = None
        edad = 0
        desde = ""
        if "desde" in columnas:
            desde = _cs(ws.cell(row_idx, columnas["desde"]).value)
        diagnostico = diagnostico_bruto
        profesional = profesional_bruto
        prioridad = prioridad_normalizada(prioridad_bruta or "")
        observaciones = observaciones_brutas
        rut = rut_raw.replace(".", "").replace("-", "").upper().strip()
        mayor_60 = False
        categoria = categoria_por_diagnostico(diagnostico, 0)

        try:
            if "fecha" in columnas:
                fecha_derivacion = _parsear_fecha(ws.cell(row_idx, columnas["fecha"]).value)
            if fecha_derivacion is None:
                raise ValueError(
                    f"Fecha inválida o ausente: {ws.cell(row_idx, columnas.get('fecha', 1)).value!r}"
                )
            fechas_hoja.append(fecha_derivacion)

            if "edad" in columnas:
                edad_valor = ws.cell(row_idx, columnas["edad"]).value
                try:
                    edad = int(float(str(edad_valor))) if edad_valor not in (None, "") else 0
                except (ValueError, TypeError):
                    edad = 0
            mayor_60 = edad >= 60
            # Si hay columna explícita ≥60, usarla para mayor_60
            if "mayor_60" in columnas:
                val_m60 = _normalizar(_cs(ws.cell(row_idx, columnas["mayor_60"]).value))
                if val_m60 in {"SI", "S", "1", "TRUE", "X"}:
                    mayor_60 = True
                elif val_m60 in {"NO", "N", "0", "FALSE", ""}:
                    mayor_60 = edad >= 60  # fallback a cálculo por edad
            categoria = categoria_por_diagnostico(diagnostico, edad)

            if not diagnostico:
                raise ValueError("Diagnóstico vacío")

            dup_key = _dup_key(rut, fecha_derivacion, diagnostico)
            if dup_key in duplicados_vistos or dup_key in existentes:
                duplicados += 1
                registros.append(_registro_preview(
                    hoja=sheet_name, fila=row_idx, nombre=nombre, rut=rut,
                    fecha_derivacion=fecha_derivacion, edad=edad,
                    diagnostico=diagnostico, prioridad=prioridad, percapita_desde=desde,
                    profesional=profesional, observaciones=observaciones,
                    mayor_60=mayor_60, categoria=categoria, es_duplicado=True,
                ))
                continue

            duplicados_vistos.add(dup_key)
            pacientes_a_crear.append(Paciente(
                id_ccr=f"TMP-{uuid.uuid4().hex[:12].upper()}",
                fecha_derivacion=fecha_derivacion,
                percapita_desde=desde,
                nombre=nombre.upper(),
                rut=rut,
                edad=edad,
                diagnostico=diagnostico,
                profesional=profesional,
                prioridad=prioridad,
                categoria=categoria,
                mayor_60=mayor_60,
                observaciones=observaciones,
            ))
            registros.append(_registro_preview(
                hoja=sheet_name, fila=row_idx, nombre=nombre, rut=rut,
                fecha_derivacion=fecha_derivacion, edad=edad,
                diagnostico=diagnostico, prioridad=prioridad, percapita_desde=desde,
                profesional=profesional, observaciones=observaciones,
                mayor_60=mayor_60, categoria=categoria,
            ))
        except Exception as exc:
            error_msg = str(exc)
            errores.append({"hoja": sheet_name, "fila": row_idx, "motivo": error_msg})
            registros.append(_registro_preview(
                hoja=sheet_name, fila=row_idx, nombre=nombre, rut=rut,
                fecha_derivacion=fecha_derivacion, edad=edad,
                diagnostico=diagnostico, prioridad=prioridad, percapita_desde=desde,
                profesional=profesional, observaciones=observaciones,
                mayor_60=mayor_60, categoria=categoria, error=error_msg,
            ))

    return {
        "registros": registros,
        "pacientes_a_crear": pacientes_a_crear,
        "errores": errores,
        "total": total,
        "duplicados": duplicados,
        "fechas_hoja": fechas_hoja,
    }


def _mes_desde_fechas(fechas: list[date]) -> str | None:
    """Detecta el mes dominante de una lista de fechas de derivación."""
    if not fechas:
        return None
    conteo: dict[int, int] = {}
    for f in fechas:
        conteo[f.month] = conteo.get(f.month, 0) + 1
    mes_num = max(conteo, key=lambda k: conteo[k])
    # Invertir MESES_NUM
    for nombre, num in MESES_NUM.items():
        if num == mes_num:
            return nombre
    return None


def _procesar_derivaciones(archivo) -> dict:
    wb = load_workbook(filename=archivo, data_only=True)

    total = 0
    duplicados = 0
    errores: list[dict] = []
    registros: list[dict] = []
    pacientes_a_crear: list[Paciente] = []
    meses_detectados: dict[str, int] = {}
    duplicados_vistos: set[tuple[str, date, str]] = set()
    meses_ignorar = {_normalizar(m) for m in MESES_IGNORAR}
    existentes = {
        _dup_key(rut, fecha_derivacion, diagnostico)
        for rut, fecha_derivacion, diagnostico in Paciente.objects.values_list(
            "rut", "fecha_derivacion", "diagnostico"
        )
    }

    for sheet_name in wb.sheetnames:
        normalizado = _normalizar(sheet_name)
        if normalizado in meses_ignorar:
            continue

        ws = wb[sheet_name]

        # Intentar detectar columnas dinámicamente
        col_info = _detectar_columnas_dinamicas(ws)
        if col_info is None:
            # Hoja sin encabezados reconocibles → saltar silenciosamente
            continue

        header_row = col_info.pop("header_row")
        columnas = col_info

        resultado_hoja = _procesar_hoja(
            ws, sheet_name, columnas, header_row,
            existentes, duplicados_vistos
        )

        total += resultado_hoja["total"]
        duplicados += resultado_hoja["duplicados"]
        errores.extend(resultado_hoja["errores"])
        registros.extend(resultado_hoja["registros"])
        pacientes_a_crear.extend(resultado_hoja["pacientes_a_crear"])
        validos_hoja = len(resultado_hoja["pacientes_a_crear"])

        # Determinar el mes de la hoja:
        # 1. Por nombre de hoja (ENERO, FEBRERO, etc.) → todos los validos van a ese mes
        # 2. Archivo base mezclado → distribuir por mes real de derivación
        mes_hoja = _mes_desde_hoja(sheet_name)
        if mes_hoja:
            # Hoja de un mes fijo: sumar todos los válidos a ese mes
            meses_detectados[mes_hoja] = meses_detectados.get(mes_hoja, 0) + validos_hoja
        else:
            # Archivo base mezclado: distribuir cada paciente por su mes real
            for f in resultado_hoja["fechas_hoja"]:
                mes_nombre = None
                for nm, num in MESES_NUM.items():
                    if num == f.month:
                        mes_nombre = nm
                        break
                if mes_nombre:
                    meses_detectados[mes_nombre] = meses_detectados.get(mes_nombre, 0) + 1

    return {
        "total": total,
        "duplicados": duplicados,
        "errores": errores,
        "registros": registros,
        "pacientes": pacientes_a_crear,
        "meses_detectados": meses_detectados,
    }


def previsualizar_derivaciones(archivo) -> dict:
    procesado = _procesar_derivaciones(archivo)
    return {
        "total": procesado["total"],
        "validos": len(procesado["pacientes"]),
        "duplicados": procesado["duplicados"],
        "errores": procesado["errores"],
        "registros": procesado["registros"],
        "meses_detectados": procesado["meses_detectados"],
    }


def parsear_derivaciones(archivo) -> dict:
    procesado = _procesar_derivaciones(archivo)
    return {
        "total": procesado["total"],
        "importados": 0,  # Will be populated by view
        "duplicados": procesado["duplicados"],
        "errores": procesado["errores"],
        "meses_detectados": procesado["meses_detectados"],
        "pacientes": procesado["pacientes"],
    }


normalizar_texto = _normalizar
parsear_fecha = _parsear_fecha
