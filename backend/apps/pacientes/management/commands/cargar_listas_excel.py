from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from apps.pacientes.models import DiagnosticoCatalogo, Paciente
from apps.importar.services import normalizar_texto


MAP_CATEGORIA = {
    "BORRADOR": Paciente.Categoria.BORRADOR,
    "MAS65": Paciente.Categoria.MAS65,
    "MAYOR O IGUAL 65": Paciente.Categoria.MAS65,
    "OA_MENOS65": Paciente.Categoria.OA_MENOS65,
    "OA MENOR 65": Paciente.Categoria.OA_MENOS65,
    "HOMBROS": Paciente.Categoria.HOMBROS,
    "LUMBAGOS": Paciente.Categoria.LUMBAGOS,
    "SDNT": Paciente.Categoria.SDNT,
    "SDT": Paciente.Categoria.SDT,
    "OTROS_NEUROS": Paciente.Categoria.OTROS_NEUROS,
    "OTROS NEUROS": Paciente.Categoria.OTROS_NEUROS,
    "AATT": Paciente.Categoria.AATT,
    "DUPLA": Paciente.Categoria.DUPLA,
}


class Command(BaseCommand):
    help = "Carga categorías y diagnósticos desde la hoja LISTAS de un Excel."

    def add_arguments(self, parser):
        parser.add_argument("archivo", type=str, help="Ruta al archivo Excel origen.")

    def handle(self, *args, **options):
        ruta = options["archivo"]
        try:
            wb = load_workbook(ruta, data_only=True)
        except Exception as exc:
            raise CommandError(f"No se pudo abrir el archivo: {exc}") from exc

        if "LISTAS" not in wb.sheetnames:
            raise CommandError("No existe la hoja LISTAS en el archivo.")

        ws = wb["LISTAS"]
        encabezado = None
        for row_idx in range(1, 16):
            row_values = [normalizar_texto(ws.cell(row=row_idx, column=col).value) for col in range(1, 20)]
            if "CATEGORIA" in row_values and ("DIAGNOSTICO" in row_values or "DIAGNOSTICO/S" in row_values):
                encabezado = (row_idx, row_values)
                break
        if not encabezado:
            raise CommandError("No se encontraron columnas CATEGORIA y DIAGNOSTICO en LISTAS.")

        header_row, header_values = encabezado
        col_categoria = header_values.index("CATEGORIA") + 1
        if "DIAGNOSTICO" in header_values:
            col_diagnostico = header_values.index("DIAGNOSTICO") + 1
        else:
            col_diagnostico = header_values.index("DIAGNOSTICO/S") + 1

        creados = 0
        for row_idx in range(header_row + 1, ws.max_row + 1):
            cat_raw = ws.cell(row=row_idx, column=col_categoria).value
            diag_raw = ws.cell(row=row_idx, column=col_diagnostico).value
            if not cat_raw or not diag_raw:
                continue

            categoria = MAP_CATEGORIA.get(normalizar_texto(cat_raw))
            if not categoria:
                continue
            diagnostico = str(diag_raw).strip()
            if not diagnostico:
                continue

            _, created = DiagnosticoCatalogo.objects.get_or_create(
                diagnostico=diagnostico,
                defaults={"categoria": categoria},
            )
            if created:
                creados += 1

        self.stdout.write(self.style.SUCCESS(f"Diagnósticos cargados: {creados}"))
